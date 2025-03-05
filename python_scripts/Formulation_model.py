# 0. Load Libraries
#!jupyter nbconvert --to script *.ipynb
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import BarChart,LineChart, Series, Reference

from datetime import date, datetime

import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import matplotlib

import os

from IPython.display import Image
# INPUT
PATH_IFCUSTOM = '//edcnas001/PUU-APP/BI REPORTING/Report Source Files/Tankenlijst/Tankenlijst_tabel - IFCUSTOM.xlsx'
PATH_GENEALOGY = '//edcnas001/PUU-APP/BI REPORTING/Report Output/Genealogy SFG/Genealogy SFG Components.xlsx'
PATH_SAP_ROUTINGS = 'C:/Users/ROOSEM03/OneDrive - Pfizer/Routings/OPEX - Recipe Resources.xlsx'
PATH_BULK_FORM_LABELS = '//edcnas004/PUU_STM2/FFP/2 - Formulatie/budget Formulatie/MASTERFILES  - NBC BUDGET - productmix & productfamilie/budget - routings - NPH analyse/budget/Masterfile Project/Product groepen - ASMT vergelijking.xlsx'

# PATH_LRVF_MODEL = '//edcnas004/PUU_OPEX/Plant Performance Data/Long Range/2022 LR/Q4 2022/LRVF Base Model - Q4 2022.xlsm'
PATH_LRVF_MODEL = "C:/Users/ROOSEM03/Pfizer/de MASTER team - Documents/SPPE Team/LRVF/2023/Q1/model-files/LRVF_Base_Q1_2023.xlsm"
PATH_FORM_PARAMETERS = 'Formulation Parameters.xlsx'

# OUTPUT
# Save this file by going to parent directory and then to 'General Data' folder
PATH_FILL_FORM_AS1 = '../General Data/FillForm_AS1.xlsx'
PATH_FILL_FORM_GENERAL = '../General Data/FillForm_General.xlsx'
# 1. Link filling item to bulk item
## 1.1 Get bulk items of AS1 -> IFCUSTOM
# Get all bulk items in AS1
bulk_items = pd.read_excel(PATH_IFCUSTOM, sheet_name = 'Report 2')
bulk_items = bulk_items[['Item', 'productomschrijving']].drop_duplicates()
bulk_items.columns = ['Bulk Item', 'Bulk Descr']
bulk_items.sort_values(by = 'Bulk Descr', inplace = True)

bulk_items.head()
## 1.2 Get BOMs of all items -> Genealogy
# Get all BOMs
genealogy = pd.read_excel(PATH_GENEALOGY)

print(f'Rows in genealogy: {len(genealogy)}')

col_gen = genealogy.columns
genealogy_fil = genealogy[col_gen[[1, 2, 10, 11, 12, 13]]][
    genealogy[col_gen[1]].str.contains("H") & 
    genealogy[col_gen[10]].str.contains("H") & 
    genealogy[col_gen[12]].str.contains("GM|KG")].drop_duplicates()

print(f'Rows in genealogy filtered on bulk items: {len(genealogy_fil)}')

genealogy_fil.to_clipboard()
genealogy_fil.head()
## 1.3 Extra: Merge bulk items with Genealogy General
# Idea 1: 1 bulk item can be used for multiple filling items. 1 Filling item can probabily not be produced by different bulk items, so these are unique.
fill_form_general =  genealogy_fil[genealogy_fil.columns[:4]]
fill_form_general.columns = ['Fill Item', 'Fill Descr', 'Bulk Item', 'Bulk Descr',]
fill_form_general = fill_form_general.sort_values(by = ['Fill Descr', 'Fill Item']).drop_duplicates()

# Idea 2: bulk item can also refer to further components (here only '10P SOL. SODIUM HYDROXIDE'?). 
# So bulk items that refer to other H-items are still in the first column
c_temp = fill_form_general.columns

#expand the bulk items with possible components
temp = fill_form_general.merge(fill_form_general, left_on = 'Bulk Item', right_on = 'Fill Item', how = 'left').drop(columns = ['Fill Item_y', 'Fill Descr_y'])
temp.columns = list(c_temp) + ['Component', 'Component Descr']

non_fill_items = []
for i in temp['Fill Item'].unique():
    if i in (temp['Bulk Item'].unique()):
        non_fill_items.append(i)

# exclude rows where the bulk item is expanded (to get solely true filling items in first column)
fill_form_general = temp[~temp['Fill Item'].isin(non_fill_items)].sort_values(by = 'Fill Item')

# load all relevant resources for formulation from the current SAP routings
resource_filter = [
'Bio – Bulk Costing',
'Asep - Bulk Costing',
'Asep - Bulk Suspension',
'Asep - Formulation CIP',
'Asep - Formulation SIP',
'Asep - Preparation Components (Processi',
'Bio - Preparatie Cell 1',
'Bio - Formulatie Cell 1 - VC1-FB1/2',
'Asep - Preparation General',
'Asep - Freeze Dry',
'Focu - Formulation Cell 1',
'Focu – Bulk Costing Cell 1 - Processing',
'S2F2 - Formulatie',
'Bio - Formulatie ALPO',
'Focu - Formulation Cell 2',
'Focu – Bulk Costing Cell 2 - Processing',
'Asep - Freeze/thaw Cryovessel',
'Bio– Form-Neisvac Cell1',
'Freeze/thaw',
'Asep - Formulation',
'Asep - Freeze/thaw PBAGS',
'Asep - Formulation Raw Mat Produced',
'Asep - Preparation Components (Cost)',
'Bio - Formulatie Cell 2 - VC2-FB3/4/5/6',
'Bio - Formulatie Cell 3 - VC3-FB7',
'Bio - Preparatie Cell 3']


SAP_routings = pd.read_excel(PATH_SAP_ROUTINGS, sheet_name = 'Budget - Recipe Resource')[
    ['Mat Nr', 'Description', 'Resource Desc', 'Recipe Group', 'MACHSU', 'LABSU', 'LABCU']]

SAP_routings = SAP_routings[
        SAP_routings['Mat Nr'].str.startswith('H') 
        & SAP_routings['Resource Desc'].isin(resource_filter) 
        & (SAP_routings['MACHSU'] > 1) 
        & (SAP_routings['LABSU'] > 1)
        ].drop_duplicates()


# merge the fill_form table with its corresponding routing
fill_form_general = fill_form_general.merge(SAP_routings, left_on = 'Fill Descr', right_on = 'Description', how = 'left').drop(
    columns = ['Mat Nr', 'Description', 'Recipe Group'])


fill_form_general_fil = fill_form_general.drop(columns=['Component', 'Component Descr'])

fill_form_general_fil.drop_duplicates(inplace = True)

bulk_form = pd.read_excel(PATH_BULK_FORM_LABELS, sheet_name = 'Samengevoegd', header = 1)[
        ['Formulatie item', 'Formulatie Product', 'Formulatie Proces label']]
bulk_form.drop_duplicates(inplace = True)

fill_form_general_fil = fill_form_general_fil.merge(bulk_form, left_on= 'Bulk Item', right_on='Formulatie item', how = 'left').drop(
                        columns = ['Formulatie item', 'Formulatie Product']).drop_duplicates()

fill_form_general_fil.to_clipboard()
fill_form_general_fil.head()
fill_form_general.drop(columns=['Component', 'Component Descr']).head()
## 1.3 Merge Bulk items with Genealogy
# Idea 1: 1 bulk item can be used for multiple filling items. 1 Filling item can probabily not be produced by different bulk items, so these are unique.
bulk_items = bulk_items.merge(genealogy_fil, left_on = 'Bulk Item', right_on ='SAP Material Number (Comp-BOM)', how = 'left')
fill_form =  bulk_items[bulk_items.columns[:4]]
fill_form.columns = ['Bulk Item', 'Bulk Descr', 'Fill Item', 'Fill Descr']
fill_form = fill_form[['Fill Item', 'Fill Descr', 'Bulk Item', 'Bulk Descr']]
fill_form = fill_form.sort_values(by = ['Fill Descr', 'Fill Item']).drop_duplicates()

# Idea 2: bulk item can also refer to further components (here only '10P SOL. SODIUM HYDROXIDE'?). 
# So bulk items that refer to other H-items are still in the first column
c_temp = fill_form.columns

#expand the bulk items with possible components
temp = fill_form.merge(fill_form, left_on = 'Bulk Item', right_on = 'Fill Item', how = 'left').drop(columns = ['Fill Item_y', 'Fill Descr_y'])
temp.columns = list(c_temp) + ['Component', 'Component Descr']

non_fill_items = []
for i in temp['Fill Item'].unique():
    if i in (temp['Bulk Item'].unique()):
        non_fill_items.append(i)

# exclude rows where the bulk item is expanded (to get solely true filling items in first column)
fill_form = temp[~temp['Fill Item'].isin(non_fill_items)].sort_values(by = 'Fill Item')
## 1.4 Merge Bulk items with current SAP Routings
# load all relevant resources for formulation from the current SAP routings
resource_filter = ['Asep - Bulk Costing']

SAP_routings = pd.read_excel(PATH_SAP_ROUTINGS, sheet_name = 'Budget - Recipe Resource')[
    ['Mat Nr', 'Description', 'Resource Desc', 'Recipe Group', 'MACHSU', 'LABSU', 'LABCU']]

SAP_routings = SAP_routings[
        SAP_routings['Mat Nr'].str.startswith('H') 
        & SAP_routings['Resource Desc'].isin(resource_filter) 
        & (SAP_routings['MACHSU'] > 1) 
        & (SAP_routings['LABSU'] > 1)
        ].drop_duplicates()


# merge the fill_form table with its corresponding routing
fill_form = fill_form.merge(SAP_routings, left_on = 'Fill Descr', right_on = 'Description', how = 'left').drop(
    columns = ['Mat Nr', 'Description', 'Resource Desc', 'Recipe Group'])

fill_form.drop_duplicates(inplace = True)

fill_form.head(5)
## 1.5 Get Specified Formulation Process Labels
#Load Form groups'
bulk_form = pd.read_excel(PATH_BULK_FORM_LABELS, sheet_name = 'Samengevoegd', header = 1)[
        ['Formulatie item', 'Formulatie Product', 'Formulatie Proces label']]
bulk_form.drop_duplicates(inplace = True)
bulk_form.head()
## 1.5 Fill_Form DF
* Filling item
* Bulk item
* Extra formulation component (mostly Sodium Hydroxide --> already included in FB6 occupation?)
* Total SAP routings per Filling Item
* Assigned Formulation Process Labels
fill_form_AS1 = fill_form.merge(bulk_form, left_on = 'Bulk Item', right_on ='Formulatie item', how = 'left').drop(columns = 'Formulatie Product')
fill_form_AS1.iloc[15:20]
fill_form_AS1.to_excel(PATH_FILL_FORM_AS1, index = False)
**The Fill_Form dataframe is not used further in the notebook, as it does only include data about items of AS1.**

- This approach needs to be elaborated and extended to all departments. 
- How to handle impact of component formulation on booth occupation?
- Evaluation of current SAP routings: complete? logical? Eg MACHSU for DP and filling item? Comparison to Masterfile Formulation and Formulation DLT analysis. 
- Further validation of assigned Formulation process labels
- Bulk items not strictly necessary in Capacity calculation, but more straightforward checking of booth occupation correctness
# 2. Load filling volumes from LRVF
volumes = pd.read_excel(PATH_LRVF_MODEL, sheet_name = 'Volumes')[
    ['Year','Profit center','Product','Item descr.', 'Item nr', 'Volume', 'Lot size', 'Yield','Line', 'SCENARIO']
]

# Data preparation: select years, commercial lot sizes
volumes = volumes[~volumes['Year'].isin([2019, 2020, 2021, 2022])]
volumes = volumes[(volumes['Lot size'] > 100)]
volumes['Volume'] = volumes['Volume'].round(1)

print(f'Scenarios in LRVF model: {volumes.SCENARIO.unique()}')

volumes.head()
## 2.1 Calculate Batches
# Filling yield is accounted for here, in the calculation of amount of batches

def round_away_zero(array):
    # function that rounds away from zero
    # Necessary to correctly sum multiple scenarios (positive and negative batches)
    return np.sign(array)* np.ceil(np.abs(array))

volumes['# Loten'] = volumes['Volume']/volumes['Lot size']*1000/volumes['Yield']*100
volumes.dropna(subset = ['# Loten'], inplace = True)
volumes['# Loten'] = round_away_zero(volumes['# Loten']).astype(int)
len(volumes[volumes['# Loten'].isnull()])
# 3. Load parameters and capacities
## 3.1 Formulation

# legend: doens't include the colours
Legend = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'Legend')[['Series name', 'Color', 'Line?', 'Sensitivity?', 'Outage?', 'Black/White?']]

# get the legend colours: dificult, same are rgb values, others a theme, others have 'tints'

# wb = load_workbook(path_form, data_only = True)
# sh = wb['Legend']

# for i in range(2,50):
#     print(sh[f'B{i}'].fill.start_color.index)
#     print(sh[f'B{i}'].fill.start_color.rgb)
#     print(sh[f'B{i}'].fill.start_color)

# -------------------------------------------------------------------------------------------------------------------------------------- 

# Filling to formulation label 
FillForm = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'FillForm')[['Product Label Filling', 'Formulation Label']].merge(
            pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'ProdItem')[['Product Label Filling', 'Filling Item Descr']])[
                ['Filling Item Descr', 'Product Label Filling', 'Formulation Label']
                ]
    
# -------------------------------------------------------------------------------------------------------------------------------------- 

# Formulation routing hours
FormRouting = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'FormRouting')[['Formulation Label', 'Mach F', 'Mach V', 'Weighed lotsize']]
%store FormRouting

# -------------------------------------------------------------------------------------------------------------------------------------- 

# Extra formulation volumes, not included in filling -> Vast aantal Spoelwater, NaOH, ethanol, MF, WFI batchen per jaar (arbitrair in FB6)
# Workaround: lot size set to 1000, then the volume counts as amount of batches

FB6_spoelwater = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'FB6')[['Year', 'Volume', 'Lot size', 'Yield', 'SCENARIO', '# Loten', 'Formulation Label']]

# -------------------------------------------------------------------------------------------------------------------------------------- 

# Allocation of formulation items to specific booths
Allocation_tab = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'Allocation')

Allocation_Theory =  Allocation_tab[['FB', 'Product Class', 'Products']]

Allocation_Matrix = Allocation_tab[list(Allocation_tab.columns)[5:29]]

# -------------------------------------------------------------------------------------------------------------------------------------- 

# Yearly capacities 
Yearly_cap_perc_FORM = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'Yearly Cap Perc FORM')[
                    ['Year','Booth', 'Formulation Label', 'Hours Form allocated %', 'SCENARIO']]

Yearly_cap_perc_FORM['Hours Form allocated %'] = Yearly_cap_perc_FORM['Hours Form allocated %']*100

Yearly_cap_hrs_FORM = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'Yearly Cap Hrs')[['Booth', 'Hrs/year 7d']]
### Formulation Matrix -> Booth Allocation
- 'V' means booth is validated for the product
- % means the product is allocated to a certain booth for % amount of the routing hours
- Allocation is done manually in the excel file 

Image('Allocation Matrix.png', width = 750)
# Number between 0-1 is % allocation
# empty / Nan is not validated
Booth = []
Formulation_Label = []
Allocation_percentage = []

for booth in Allocation_Matrix.columns[1:]:
    for i in range(len(Allocation_Matrix)):
        if Allocation_Matrix[booth].iloc[i] != 'V':
            if Allocation_Matrix[booth].iloc[i] > 0:
                Booth.append(booth)
                Formulation_Label.append(Allocation_Matrix['Allocation'].iloc[i])
                Allocation_percentage.append(Allocation_Matrix[booth].iloc[i])
                
Allocation = pd.DataFrame(list(zip(Booth, Formulation_Label, Allocation_percentage)), columns = ['Booth', 'Formulation Label', 'Allocation [%]'])
Allocation['Allocation [%]'] = (Allocation['Allocation [%]']*100).astype(int)

%store Allocation_Matrix


Allocation.head()
# Get all validated products per booth out of FORM ASMT-4727
# Allocation_Theory['Products'].str.split(", ", expand = True).dropna(subset = [1])

# Allocation_Theory.dropna(how = 'all', inplace = True)
# Allocation_Theory['FB'] = Allocation_Theory['FB'].fillna(method='pad').astype(int)
# Allocation_Theory = Allocation_Theory.join(Allocation_Theory['Products'].str.split(", ", expand = True))

# l = []
# for c in Allocation_Theory[Allocation_Theory.columns[3:]]:
#     l += Allocation_Theory[c].values.tolist()

# pd.DataFrame(set(l)).sort_values(by = 0)
## 3.2 Weighing
WeighRouting = pd.read_excel(PATH_FORM_PARAMETERS,
                sheet_name = 'WeighRouting')[['Formulation Label', '1st run DFB', '1st run SOL', '1st run ISO', '2nd run DFB', '2nd run SOL', '2nd run ISO', '2nd run FORM', 'Mach F WC']]

# fill in blank values with the mean of the column. This allows further estimation of the resource occupation.
WeighRouting.fillna(WeighRouting.mean(), inplace = True)

%store WeighRouting

# Yearly capacities 
Yearly_cap_perc_WEIGH = pd.read_excel(PATH_FORM_PARAMETERS,
                sheet_name = 'Yearly Cap Perc WEIGH')[['Year','Booth', 'Formulation Label', 'Hours Weigh allocated %', 'SCENARIO']]

Yearly_cap_perc_WEIGH['Hours Weigh allocated %'] = Yearly_cap_perc_WEIGH['Hours Weigh allocated %']*100

## 3.3 Vessels + CIPSIP 
VesselRouting = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'VesselRouting')[['Vessel Label', 'Vessel Type']]

%store VesselRouting

CIPSIPRouting = pd.read_excel(PATH_FORM_PARAMETERS, sheet_name = 'VesselRouting')[['Vessel', 'CIP [h]', 'SIP [h]', 'CIPSIP [h]']]
# 4. Create Model
# Add the formulation labels to the filling product volumes

model = volumes.merge(FillForm[['Product Label Filling', 'Formulation Label']], 
        left_on = 'Product', right_on ='Product Label Filling', how='left').drop(
        columns = ['Product Label Filling'])

# Append the extra non-commercial occupation in the booths
model_FORM = pd.concat([model, FB6_spoelwater])

print(len(model_FORM))
model_FORM = model_FORM.merge(FormRouting, on = 'Formulation Label', how = 'left')

print(len(model_FORM))
model_FORM = model_FORM.merge(SAP_routings.drop_duplicates(subset = ['Description']), left_on = 'Item descr.', right_on = 'Description', how = 'left')

print(len(model_FORM))

# Check that the volumes don't duplicate: size of df remains the same

model_FORM.head()
#model_FORM[model_FORM['Profit center'] == 'SOMATROGON']
## 4.1 Formulation model
FillForm = FillForm.merge(FormRouting, on = 'Formulation Label', how = 'left')
FillForm = FillForm.sort_values(by = 'Filling Item Descr').merge(SAP_routings, 
            left_on = 'Filling Item Descr', right_on = 'Description', how = 'left')
FillForm.drop_duplicates(subset = ['Product Label Filling', 'Formulation Label'], inplace = True)

FillForm.to_excel(PATH_FILL_FORM_GENERAL, index = False)

%store FillForm
FillForm.head()
FillForm[FillForm['Formulation Label'] == 'SOMATROGON']
# Assigning bulks based on filling item doesn't work: too many filling items in volume file do not correspond to BOM genealogy file.
# volumes.merge(fill_form, left_on = 'Item descr.', right_on ='Fill Descr', how = 'left').drop(columns = ['Fill Item', 'Fill Descr']).drop_duplicates().head(10)
# Check if all filling products are linked in formulation
print(len(model_FORM[model_FORM['Formulation Label'].isnull()]))
pd.DataFrame(model_FORM[model_FORM['Formulation Label'].isnull()][['Product', 'Item descr.']]).drop_duplicates()

print(model_FORM[model_FORM['Mach F'].isnull()]['Product'].unique())
## 4.2 Weighing Model
WeighRouting
model_WEIGH = model.merge(WeighRouting, how='left')
model_WEIGH.head()
# Check for products with no routing
model_WEIGH[model_WEIGH['Mach F WC'].isnull()]['Product'].unique()
for process in ['1st run DFB', '1st run SOL', '1st run ISO', '2nd run DFB', '2nd run SOL', '2nd run ISO', '2nd run FORM']:
    model_WEIGH['Hrs ' + process] = (model_WEIGH[process])*model_WEIGH["# Loten"]


model_WEIGH['Hours Weighing'] = (model_WEIGH['Mach F WC'])*model_WEIGH["# Loten"]
model_WEIGH.head()
model_WEIGH.to_clipboard()
## 4.3 CIPSIP Model


## 4.4 Formulation Allocation
len(Allocation['Formulation Label'].unique())
print(model_FORM['# Loten'].sum())
model_FORM = model_FORM.merge(Allocation, how='left', on = 'Formulation Label')
model_FORM['# Loten'] = model_FORM['# Loten'] * model_FORM['Allocation [%]'] / 100

print(model_FORM['# Loten'].sum())

# round to ceiling to get integer number of batches per row
model_FORM['# Loten'] = model_FORM['# Loten'].apply(np.ceil).astype(int)
model_FORM['Volume'] = model_FORM['Volume'].apply(np.ceil).astype(int)

# Get rid of rows with zero batches
model_FORM['# Loten'] = model_FORM['# Loten'].fillna(0).astype(int)
model_FORM.head()

print(model_FORM['# Loten'].sum())
### In Hours
# Calculate formulation time based on fixed and lot size depended duration
model_FORM['Hours Formulation'] = (model_FORM['Mach F'] + model_FORM['Mach V']*model_FORM['Weighed lotsize'])*model_FORM["# Loten"]

# Compare to SAP routings. Watch out: drop duplicates based on product (eg diluent) to avoid dubble counts
model_FORM['Hours Formulation SAP'] = (model_FORM['MACHSU'])*model_FORM["# Loten"]


# These bulk items have no formulation routings:
print('These bulk items have no formulation routings:')
print(len(model_FORM[model_FORM['Hours Formulation'].isnull()]))

# These bulk items have no SAP routings:
print('These bulk items have no SAP routings:')
print(len(model_FORM[model_FORM['Hours Formulation SAP'].isnull()]))


print(model_FORM['# Loten'].sum())

model_FORM.head()
### In Percentage
model_FORM = model_FORM.merge(Yearly_cap_hrs_FORM, on = 'Booth', how = 'left')
# Loten zijn al per booth verdeeld: niet nodig om nog eens te vermenigvuldigen met de allocatie %
# model_FORM['Hours Form allocated %'] = model_FORM['Hours Formulation']/model_FORM['Hrs/year 7d']*model_FORM['Allocation [%]']
model_FORM['Hours Form allocated %'] = model_FORM['Hours Formulation']/model_FORM['Hrs/year 7d']*100


print(model_FORM['# Loten'].sum())
model_FORM.head()
# Compare to SAP routings. Watch out: drop duplicates based on product (eg diluent) to avoid dubble counts
# Loten zijn al per booth verdeeld: niet nodig om nog eens te vermenigvuldigen met de allocatie %
# model_FORM['Hours Form allocated % SAP'] = model_FORM['Hours Formulation SAP']/model_FORM['Hrs/year 7d']*model_FORM['Allocation [%]']
model_FORM['Hours Form allocated % SAP'] = model_FORM['Hours Formulation SAP']/model_FORM['Hrs/year 7d']*100

print(model_FORM['# Loten'].sum())
model_FORM.head()
model_FORM[model_FORM['Formulation Label'] == 'Adalimumab']
# To do: Allocate the labour for weighing to appropriate resource

# model_WEIGH = model_WEIGH.merge(Yearly_cap_hrs_FORM, on = 'Booth', how = 'left')
# model_WEIGH['Hours Weigh allocated %'] = model_WEIGH['Hours Weighing'].astype(float)/model_WEIGH['Hrs/year 7d']*100
# model_WEIGH.head()
# keep the capacity lines as a different pandas dataframe: don't add them to to the model 

# for c in model_FORM.columns.tolist():
#     if c not in Yearly_cap_perc_FORM.columns.tolist():
#         Yearly_cap_perc_FORM[c] = None
# Yearly_cap_perc_FORM.head()
# model_FORM = model_FORM.append(Yearly_cap_perc_FORM)[model_FORM.columns.tolist()]
# len(model_FORM)
#model_FORM[model_FORM['Booth'] == 'AS1 FB6'][['Year', 'Formulation Label', 'Hours allocated %']].plot.bar(index = "Year")

# AS = ['AS1 FB5', 'AS1 FB6', 'AS1 FB7', 'AS1 FB8', 'AS1 FB9', 'AS1 FB10', 'AS1 FB11', 'AS1 FB12', 'Transfer']
# FC = ['Focus']
# VC = ['VC FB1/2/7', 'VC FB3-6-8']

# for SCEN in model_FORM['SCENARIO'].unique():
#     for afd in [AS, FC, VC]:
#         print('Scenario: ' + SCEN)
#         display(pd.pivot_table(model_FORM[(model_FORM['Correct allocation'] == True) & (model_FORM['Scenario allocation'] == 'Base Case') & (model_FORM['SCENARIO'].isin(['Base Case', SCEN])) & (model_FORM['Booth'].isin(afd))], 
#                                index = ['Year'], columns=['Booth','Formulation Label'], values = ['# Loten'], aggfunc=np.sum, fill_value = '0', margins = True))
## 4.2 Show data of interest
model_FORM = model_FORM[~model_FORM['Year'].isin([2019, 2020, 2021, 2022])]
col_basic = ['Year', 'Profit center', 'Product', 'Item descr.', 'Volume', 'Lot size', '# Loten', 'SCENARIO', 'Formulation Label']
col_form = ['Mach F', 'Booth', 'Hours Formulation', 'Hours Form allocated %']
col_weigh = ['1st run DFB', '1st run SOL', '1st run ISO', '2nd run DFB', '2nd run SOL', '2nd run ISO', '2nd run FORM', 'Mach F WC', 'Hours Weigh allocated %']
## 4.3 Store model
%store model_FORM
print(len(model_FORM))
model_FORM.drop_duplicates(inplace=True)
print(len(model_FORM))
today = date.today()
d1 = today.strftime("%d-%m-%Y")

model_FORM.to_excel('Formulatie_model_'+d1+'.xlsx', index = False)
# 5 Capacity Plots
def addlabels(x,labels_loten, labels_occ):
    
    for i in range(len(x)):
        for j in range(1, len(labels_loten[i])):
            if float(labels_loten[i][j]) > 15:
                
                
                plt.text(i, sum([float(k) for k in labels_occ[i][1:j]]) + float(labels_occ[i][j])//2, int(labels_loten[i][j]), ha = 'center')



def cap_plot(df, resource, x_axis = 'Year', values = 'Hours Form allocated %', scenario = ['Base Case'], label = 'Formulation Label'):
    
    years_to_show = list(range(2023,2031))
    df = df.loc[df['Year'].isin(years_to_show)].copy()
    
    df['# Loten'] = df['# Loten'].astype(float)
    df_filt = df[(df['Booth'] == resource) & (df['SCENARIO'].isin(scenario))][[x_axis, values, label]]
    df_labels_filt = df[(df['Booth'] == resource) & (df['SCENARIO'].isin(scenario))][[x_axis, '# Loten', label]]
    df_volume_filt = df[(df['Booth'] == resource) & (df['SCENARIO'].isin(scenario))][[x_axis, 'Volume', label]]
    
    
    
    # dataframe for the plot needs to have the years in the first column (index), and the product labels in the next columns. The values are summed per year for the labels.
    # the capacity lines should be in a different dataframe and superimposed on the bar chart
    
    df_plot = df_filt.pivot_table(index = [x_axis], columns = [label], values = values, aggfunc = np.sum, fill_value = 0).reset_index()
    df_plot[x_axis] = df_plot[x_axis].astype('str')
    
    df_labels = df_labels_filt.pivot_table(index = [x_axis], columns = [label], values = '# Loten', aggfunc = np.sum, fill_value = 0).reset_index()
    df_labels[x_axis] = df_labels[x_axis].astype('str')
    
    df_volume = df_volume_filt.pivot_table(index = [x_axis], columns = [label], values = 'Volume', aggfunc = np.sum, fill_value = 0).reset_index()
    df_volume[x_axis] = df_volume[x_axis].astype('str')
    
    #df_labels['# Loten'] = df_labels['# Loten'].astype(float)
    
    # sort the values based on the sum of all rows of a column
    # Create the dataframe with all sums, reorder by descending totals, and collect this column sequence
    
    temp_df = df_filt.pivot_table(index = [x_axis], columns = [label], values = values, aggfunc = np.sum, fill_value = 0, margins = True).sort_values(
        'All', ascending=False, axis=1).drop('All', axis = 1).drop('All', axis = 0).reset_index()

    # column order to be used in the graph
    col_order = temp_df.columns
    
    # To do: get colours out of legend to show in the graph
    
    #...
    
    
    
    #set style and color cycle
    matplotlib.style.use('seaborn-talk')
    cm = plt.get_cmap('tab20')
    matplotlib.rcParams['axes.prop_cycle'] = matplotlib.cycler(color= [cm(1.*i/20) for i in range(20)]) 
    
    bar_width = 0.5
    ax = df_plot[col_order].plot(x=x_axis, kind='bar',  stacked=True, width = bar_width, ylabel = 'Occupation [%]')
    #ax = df_ssc_plot.plot(x= x_axis ,ax=ax)
    
    # get # Loten per formulation group as label on te graph (instead of % occupation)
    addlabels(df_labels[x_axis], df_labels[col_order].values.tolist(), df_plot[col_order].values.tolist()) 
    
    
    
#     for c in ax.containers:
        
#         # Optional: if the segment is small or 5, customize the labels
#         labels = [round(v.get_height(),1) if v.get_height() > 5 else '' for v in c]

#         # remove the labels parameter if it's not needed for customized labels
#         ax.bar_label(c, labels = labels, label_type='center', fmt='%.1f')
    
    
    df_plot['Cap 7d%'] = 70
    df_plot['Cap 6d%'] = 60
    df_plot['Cap 5d%'] = 50
    df_plot[['Year', 'Cap 7d%', 'Cap 6d%', 'Cap 5d%']].plot(x=x_axis, kind='line', ax = ax, 
                                                            style = {'Cap 7d%': 'r', 'Cap 6d%' : 'b', 'Cap 5d%': 'g'})
    

    
    
    # Put a legend to the right of the current axis
    # Order the legend names as the stacked bars
    
    leg_handles, leg_labels = plt.gca().get_legend_handles_labels()
    order = [0,1,2] + list(range(len(leg_labels)-1, 2, -1))
    
    ax.legend([leg_handles[idx] for idx in order],[leg_labels[idx] + ' (' + str(FormRouting_graph.loc[FormRouting_graph['Formulation Label'] == leg_labels[idx], 'Mach F'].item()) + 'h)' for idx in order], loc='center left', bbox_to_anchor=(1, 0.5))
    ax.grid(False)
    
    
    # Assign a title
    new_line = '\n'
    
    scenario_str = ' + '.join(scenario)
    title = f"{resource} : {values} and Batches per year {new_line} Scenario: {scenario_str}"
    plt.title(label = title)
    
    plt.show()
    
#     display(df_labels.merge(df_volume, on = 'Year', suffixes = (' [Bulk]', ' [\'000]')))
    display(df_labels.set_index('Year'))
    display(df_volume.set_index('Year'))
    
    
    return df_plot
## Calculated Routings
FormRouting_graph = FormRouting.append(pd.DataFrame([['Cap 7d%', '', '',''], ['Cap 6d%', '', '',''], ['Cap 5d%', '', '','']], columns = FormRouting.columns)).reset_index()
AS1_booths = ['AS1 FB5', 'AS1 FB6', 'AS1 FB7', 'AS1 FB8', 'AS1 FB9', 'AS1 FB10', 'AS1 FB11', 'AS1 FB12']
AS1_extra = ['AS1 Filtration', 'AS1 Transfer', 'AS1 Suspension']
VC_booths = ['VC FB1', 'VC FB2', 'VC FB3', 'VC FB4', 'VC FB5', 'VC FB6', 'VC FB7', 'VC FB8', 'VC ALPO']
FC_booths = ['Focus FB1', 'Focus FB1']


for booth in (AS1_booths):
    display(cap_plot(df = model_FORM, resource = booth, scenario = ['Base Case', 'LE']).set_index('Year'))
## SAP routings
for booth in (AS1_booths):
    display(cap_plot(df = model_FORM[['Year', 'Profit center', 'Booth', 'SCENARIO', '# Loten', 'Hours Form allocated % SAP']].drop_duplicates(), 
                     resource = booth, values = 'Hours Form allocated % SAP', label = 'Profit center'))
    
## 5.1 Validation
### Total amount of batches per year per booth
def batches_year(df, booths, x_axis = 'Year', values = '# Loten', scenario = ['Base Case'], label = 'Formulation Label'):
    
    df['# Loten'] = df['# Loten'].astype(float)
    df_filt = df[(df['Booth'].isin(booths)) & (df['SCENARIO'].isin(scenario))][[x_axis, 'Booth', values, label]]    
    
    df_batches = df_filt.pivot_table(index = [x_axis], columns = ['Booth'], values = values, aggfunc = np.sum, fill_value = 0, margins = True).sort_values(
        'All', ascending=False, axis=1).drop('All', axis = 0).reset_index()
    
    return df_batches
batches_year(df=model_FORM, booths = (AS1_booths))
## 5.x Extra code
(AS1_booths + AS1_extra)
def cap_plot(df, resource, x_axis = 'Year', values = 'Hours Form allocated %', scenario = ['Base Case'], label = 'Formulation Label'):
    
    df_filt = df[(df['Booth'] == resource) & (df['SCENARIO'].isin(scenario))][[x_axis, values, label]]
    df_labels_filt = df[(df['Booth'] == resource) & (df['SCENARIO'].isin(scenario))][[x_axis, '# Loten', label]]
    
    # dataframe for the plot needs to have the years in the first column (index), and the product labels in the next columns. The values are summed per year for the labels.
    # the capacity lines should be in a different dataframe and superimposed on the bar chart
    
    df_plot = df_filt.pivot_table(index = [x_axis], columns = [label], aggfunc = np.sum, fill_value = 0).reset_index()
    df_plot[x_axis] = df_plot[x_axis].astype('str')
    
    df_labels = df_labels_filt.pivot_table(index = [x_axis], columns = [label], aggfunc = np.sum, fill_value = 0).reset_index()
    df_labels[x_axis] = df_labels[x_axis].astype('str')
    
    
    # To do: sort the values based on the sum of all rows of a column 
    
    # ...
    
    
    # To do: get # Loten per formulation group as label on te graph (instead of % occupation)
    
    # ...
    
    
    # To do: get colours out of legend to show in the graph
    
    #...
    
    
    
    #set style and color cycle
    matplotlib.style.use('seaborn-talk')
    cm = plt.get_cmap('tab20')
    matplotlib.rcParams['axes.prop_cycle'] = matplotlib.cycler(color= [cm(1.*i/20) for i in range(20)]) 
    
    bar_width = 0.5
    ax = df_plot.plot(x=x_axis, kind='bar',  stacked=True, width = bar_width, ylabel = 'Occupation [%]')
    #ax = df_ssc_plot.plot(x= x_axis ,ax=ax)
    
    for c in ax.containers:
        
        # Optional: if the segment is small or 5, customize the labels
        labels = [round(v.get_height(),1) if v.get_height() > 5 else '' for v in c]

        # remove the labels parameter if it's not needed for customized labels
        ax.bar_label(c, labels = labels, label_type='center', fmt='%.1f')
    
    
    df_plot['Cap 7d%'] = 70
    df_plot['Cap 6d%'] = 60
    df_plot['Cap 5d%'] = 50
    df_plot[['Year', 'Cap 7d%', 'Cap 6d%', 'Cap 5d%']].plot(x=x_axis, kind='line', ax = ax)
    
    
    # Put a legend to the right of the current axis
    ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    ax.grid(False)
    
    # Assign a title
    scenario_str = ' - '.join(scenario)
    title = f'{resource} : {values} in scenario {scenario_str}'
    plt.title(label = title)
    
    plt.show()
    
    return df_plot
# 6 Export Data Frame to Excel
today = date.today()
d1 = today.strftime("%d-%m-%Y")

new_workbook = f"Formulatie_model_Q1_2022_{d1}.xlsx"

with pd.ExcelWriter(new_workbook, engine='xlsxwriter') as writer:
    workbook = writer.book
    # workbook.add_worksheet('Legend')

    # Convert the dataframe to an XlsxWriter Excel object. Turn off the default
    # header and index and skip one row to allow us to insert a user defined
    model_FORM.to_excel(writer, sheet_name = 'model', startrow=1, header= False, index = False, engine = 'xlsxwriter')

    worksheet = writer.sheets['model']

    # Get the dimensions of the dataframe.
    (max_row, max_col) = model_FORM.shape

    # Create a list of column headers, to use in add_table().
    column_settings = []
    for header in model_FORM.columns:
        column_settings.append({'header': header})

    # Add the table.
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

    # Make the columns wider for clarity.
    worksheet.set_column(0, max_col - 1, 12)

    workbook.add_worksheet('Pivot')

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()
import win32com.client
Excel   = win32com.client.gencache.EnsureDispatch('Excel.Application') # Excel = win32com.client.Dispatch('Excel.Application')

win32c = win32com.client.constants
#disable pop-ups
Excel.Application.DisplayAlerts = False

#load form Q4 model
wb_gen_model = Excel.Workbooks.Open(path_form, ReadOnly = 1)

#open newly created workbook
wb = Excel.Workbooks.Open(f'C:/Users/ROOSEM03/Jupyter Lab/Formulatie Model/{new_workbook}')

#Haal legend uit formulatie model!
Sheet0 = wb_gen_model.Worksheets["Legend"]
Sheet0.Copy(Before = wb.Worksheets(1))

wb_gen_model.Close()

Sheet1 = wb.Worksheets["model"]
Sheet1.Select()

cl1 = Sheet1.Cells(1,1)
cl2 = Sheet1.Cells(len(model_FORM),len(model_FORM.columns))
PivotSourceRange = Sheet1.Range(cl1,cl2)

PivotSourceRange.Select()

Pivot_sheet = wb.Worksheets['Pivot']

row = 50
dist = 25

for i, booth in zip(['1', '2', '3', '4'], ['AS1 FB6', 'AS1 FB7', 'AS1 FB8', 'AS1 FB9']):
    for name in ['pvtgraph', 'pvtvolume']: 
        
        print('Start: ' + name + i + ' ' + booth)
        
        cl3=Pivot_sheet.Cells(row,1)
        PivotTargetRange=  Pivot_sheet.Range(cl3,cl3)
        PivotTableName = name + i

        PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)
        PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)

        PivotTable.PivotFields('Year').Orientation = win32c.xlRowField
        PivotTable.PivotFields('Year').Position = 1
        PivotTable.PivotFields('Booth').Orientation = win32c.xlPageField
        PivotTable.PivotFields('Booth').Position = 1
        PivotTable.PivotFields('SCENARIO').Orientation = win32c.xlPageField
        PivotTable.PivotFields('SCENARIO').Position = 2
        PivotTable.PivotFields('SCENARIO').CurrentPage = 'Base Case'
        PivotTable.PivotFields('Profit center').Orientation = win32c.xlPageField
        PivotTable.PivotFields('Profit center').Position = 3
#         PivotTable.PivotFields('Scenario allocation').Orientation = win32c.xlPageField
#         PivotTable.PivotFields('Scenario allocation').Position = 4
#         PivotTable.PivotFields('Scenario allocation').CurrentPage = 'Base Case'
#         PivotTable.PivotFields('Correct allocation').Orientation = win32c.xlPageField
#         PivotTable.PivotFields('Correct allocation').Position = 4
#         PivotTable.PivotFields('Correct allocation').CurrentPage = 'True'
        PivotTable.PivotFields('Booth').CurrentPage = booth
        PivotTable.PivotFields('Formulation Label').Orientation = win32c.xlColumnField
        PivotTable.PivotFields('Formulation Label').Position = 1
        PivotTable.PivotFields('Formulation Label').Subtotals = [False, False, False, False, False, False, False, False, False, False, False, False]
        # PivotTable.PivotFields('SCENARIO').Orientation = win32c.xlColumnField
        # PivotTable.PivotFields('SCENARIO').Position = 2
        
        PivotTable.FieldListSortAscending = True

        if name == 'pvtgraph':

            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Hours Form allocated %'))
            # DataField.NumberFormat = '###0.00'
            DataField.NumberFormat = '###0.0'
            

        elif name == 'pvtvolume':

            DataField = PivotTable.AddDataField(PivotTable.PivotFields('# Loten'))
            # DataField.NumberFormat = '###0.00'
            DataField.NumberFormat = '###0'
            
#         chart=Pivot_sheet.Shapes.AddChart2(201) 

        row += dist
        print('End: ' + name + i + ' ' + booth)
        
# total volume
cl3=Pivot_sheet.Cells(50,20)
PivotTargetRange=  Pivot_sheet.Range(cl3,cl3)
PivotTableName = 'Total volume'

PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)
PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)

PivotTable.PivotFields('Year').Orientation = win32c.xlRowField
PivotTable.PivotFields('Year').Position = 1
PivotTable.PivotFields('Booth').Orientation = win32c.xlPageField
PivotTable.PivotFields('Booth').Position = 1
PivotTable.PivotFields('SCENARIO').Orientation = win32c.xlPageField
PivotTable.PivotFields('SCENARIO').Position = 2
PivotTable.PivotFields('SCENARIO').CurrentPage = 'Base Case'
PivotTable.PivotFields('Profit center').Orientation = win32c.xlPageField
PivotTable.PivotFields('Profit center').Position = 3
# PivotTable.PivotFields('Scenario allocation').Orientation = win32c.xlPageField
# PivotTable.PivotFields('Scenario allocation').Position = 4
# PivotTable.PivotFields('Scenario allocation').CurrentPage = 'Base Case'
PivotTable.PivotFields('Booth').Orientation = win32c.xlColumnField
PivotTable.PivotFields('Booth').Position = 1
PivotTable.PivotFields('Booth').Subtotals = [False, False, False, False, False, False, False, False, False, False, False, False]
PivotTable.PivotFields('Formulation Label').Orientation = win32c.xlColumnField
PivotTable.PivotFields('Formulation Label').Position = 2
PivotTable.PivotFields('Formulation Label').Subtotals = [False, False, False, False, False, False, False, False, False, False, False, False]

DataField = PivotTable.AddDataField(PivotTable.PivotFields('# Loten'))
# DataField.NumberFormat = '###0.00'
DataField.NumberFormat = '###0'

Excel.Visible = 1

# run a macro saved in the file:
# xl.Application.Run("excelsheet.xlsm!modulename.macroname")

wb.Worksheets['Pivot'].Select()
print('Macros')
# Excel.Application.Run("C:/Users/ROOSEM03/Jupyter Lab/Formulatie Model/Formulatie_model.xlsm!Sheet2.Macro2")
# Excel.Application.Run("C:/Users/ROOSEM03/Jupyter Lab/Formulatie Model/Formulatie_model.xlsm!Sheet2.verandervolgorde")
# Excel.Application.Run("C:/Users/ROOSEM03/Jupyter Lab/Formulatie Model/Formulatie_model.xlsm!Sheet2.Haalkleurop")

print('Done.')

# wb.SaveAs('C:/Users/ROOSEM03/Jupyter Lab/Formulatie Model/Formulatie_model.xlsx')
wb.Save()

import win32com.client

Excel.Workbooks.Open(Filename="C:/Users/ROOSEM03/Jupyter Lab/Formulatie Model/Formulatie_model_new_Q1_2022.xlsx)
Excel.Workbooks.Open(Filename="C:/Users/ROOSEM03/Jupyter Lab/Formulatie Model/Formulatie_model.xlsm",ReadOnly=1)
xl.Application.Run("macro")
xl = 0
import os
import sys

# Import System libraries
import glob
import random
import re

sys.coinit_flags = 0 # comtypes.COINIT_MULTITHREADED

# USE COMTYPES OR WIN32COM
#import comtypes
#from comtypes.client import CreateObject

# USE COMTYPES OR WIN32COM
import win32com
from win32com.client import Dispatch

scripts_dir = "C:\\scripts"
conv_scripts_dir = "C:\\converted_scripts"
strcode = \
'''
sub test()
   msgbox "Inside the macro"
end sub
'''

#com_instance = CreateObject("Excel.Application", dynamic = True) # USING COMTYPES
com_instance = Dispatch("Excel.Application") # USING WIN32COM
com_instance.Visible = True 
com_instance.DisplayAlerts = False 

for script_file in glob.glob(os.path.join(scripts_dir, "*.xls")):
    print "Processing: %s" % script_file
    (file_path, file_name) = os.path.split(script_file)
    objworkbook = com_instance.Workbooks.Open(script_file)
    xlmodule = objworkbook.VBProject.VBComponents.Add(1)
    xlmodule.CodeModule.AddFromString(strcode.strip())
    objworkbook.SaveAs(os.path.join(conv_scripts_dir, file_name))

com_instance.Quit()
import xlwings as xw
xw.view(model)
import win32com.client as win32
win32c = win32.constants
import sys
import itertools
tablecount = itertools.count(1)

def addpivot(wb,sourcedata,title,filters=(),columns=(),
         rows=(),sumvalue=(),sortfield=""):

    newsheet = wb.Sheets.Add()
    newsheet.Cells(1,1).Value = title
    newsheet.Cells(1,1).Font.Size = 16
    tname = "PivotTable%d"%tablecount.next()
    pc = wb.PivotCaches().Add(SourceType=win32c.xlDatabase,
                             SourceData=sourcedata)
    pt = pc.CreatePivotTable(TableDestination="%s!R4C1"%newsheet.Name,
                         TableName=tname,
                         DefaultVersion=win32c.xlPivotTableVersion10)
    for fieldlist,fieldc in ((filters,win32c.xlPageField),
                        (columns,win32c.xlColumnField),
                        (rows,win32c.xlRowField)):
        for i,val in enumerate(fieldlist):
            wb.ActiveSheet.PivotTables(tname).PivotFields(val).Orientation = fieldc
            wb.ActiveSheet.PivotTables(tname).PivotFields(val).Position = i+1
    wb.ActiveSheet.PivotTables(tname).AddDataField(wb.ActiveSheet.PivotTables(tname).
                                         PivotFields(sumvalue),sumvalue,win32c.xlSum)
# Go to Excel
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import BarChart,LineChart, Series, Reference
from copy import deepcopy
import xlsxwriter
### Option 1: Create new file with data in sheet
#List of column name dictionaries
headers=[{"header" : i} for i in list(model.columns)]

#Create and propagate workbook
workbook=xlsxwriter.Workbook('output.xlsx')

worksheet1=workbook.add_worksheet()

worksheet1.add_table(0, 0, len(model), len(model.columns)-1, {"columns":headers, "data":model.values.tolist()})

workbook.close()
### Option 2: Create new file
- legend LRVF base model
- pandas data
- pivot tables
# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter('Model_2.xlsx', engine='xlsxwriter')

# Convert the dataframe to an XlsxWriter Excel object. Turn off the default
# header and index and skip one row to allow us to insert a user defined
model.to_excel('Model_2.xlsx', sheet_name = 'Sheet1', startrow = 1, header = False,  index = False)

# Get the xlsxwriter workbook and worksheet objects.
workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Get the dimensions of the dataframe.
(max_row, max_col) = model.shape

# Create a list of column headers, to use in add_table().
column_settings = []
for header in model.columns:
    column_settings.append({'header': header})

# Add the table.
worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

# Make the columns wider for clarity.
worksheet.set_column(0, max_col - 1, 12)

# Close the Pandas Excel writer and output the Excel file.
writer.save()

wb = Workbook(write_only = True)
ws = wb.create_sheet()

for r in dataframe_to_rows(pd.pivot_table(model[model['Booth'] == 'AS1 FB10'], index = ['Year'], columns='Formulation Label', values = ['Hours allocated %'], aggfunc=np.sum, fill_value = '0', margins = True), index  = True, header = True):
    ws.append(r)
    
chart1 = BarChart()
chart1.style = 10
chart1.type = "col"
chart1.grouping = "stacked"
chart1.overlap = 100
chart1.title = "Stacked Chart"

data = Reference(ws, min_col=2, min_row=2, max_row=13, max_col=8)
cats = Reference(ws, min_col=1, min_row=4, max_row=13)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 1
chart1.y_axis.title = 'Perc'
chart1.y_axis.majorGridlines = None


for r in dataframe_to_rows(pd.pivot_table(Yearly_capacities_perc[Yearly_capacities_perc['Booth'] == 'AS1 FB10'], index = ['Year'], columns='Formulation Label', values = ['Hours allocated %'], aggfunc=np.sum, fill_value = '0', margins = True), index  = True, header = True):
    ws.append(r)

chart2 = LineChart()
data2 = Reference(ws, min_col=2, min_row=15, max_row=26, max_col=7)
chart2.add_data(data, titles_from_data=True)
chart2.y_axis.axId = 200
chart2.y_axis.title = "Perc"



chart1.y_axis.crosses = "max"



chart1 += chart2

ws.add_chart(chart1, "H2")

wb.save("col.xlsx")
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
%matplotlib inline
test = pd.pivot_table(model[model['Booth'] == 'AS1 FB10'], index = ['Year'], columns='Formulation Label', values = ['Hours allocated %'], aggfunc=np.sum, fill_value = '0', margins = True)
test.index
pd.pivot_table(Yearly_capacities_perc[Yearly_capacities_perc['Booth'] == 'AS1 FB10'], index = ['Year'], columns='Formulation Label', values = ['Hours allocated %'], aggfunc=np.sum, fill_value = '0', margins = True)
# Pivot table
